#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/5/18 17:45
# @Author  : Puxf
# @Site    : 
# @File    : excel_test.py
# @Software: PyCharm


import openpyxl

#创建一个工作簿
wb = openpyxl.workbook()
#创建一个test_case的表单
wb.create_sheet('test_case')
#保存为一个xlsx格式的文件
wb.save('cases.xlsx')
#读取excel中的数据
#第一步打开工作簿
wb = openpyxl.load_workbook('cases.xlsx')
#第二步：选取表单
sh = wb['Sheet1']
#参数 row:行  column;列
ce = sh.cell(row = 1,column=1)
print(ce.value)
#按行读取数据 list(sh.rows)
print(list(sh.rows)[1:1])
for  cases in (list(sh.rows)[1:]):
    case_id = cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_excepted,case_data)
#关闭工作簿
wb.close()
